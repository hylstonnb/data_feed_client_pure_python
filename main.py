import datetime
import logging
import os
import threading
import time

import openpyxl
import pytz

import utils
import pandas as pd

log_file_name = f"{datetime.datetime.now().strftime('%Y-%m-%d')}"
log_file_path = utils.get_dir_path('logs\\' + log_file_name)
logging_main_handler = logging.FileHandler(log_file_path + '.log')
logging_main_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
logging_main_handler.setFormatter(formatter)
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
logger.addHandler(logging_main_handler)
loop_duration = 60
stocks_list = {}
tmzone = pytz.timezone("America/SAO_PAULO")
trade_data_columns = ['trade_date', 'trade_number', 'price', 'amount', 'volume', 'buy_agent', 'sell_agent',
                      'trade_type']
lock = threading.Lock()
import nelogica_data_feed_api


def main():
    try:
        logger.info('Starting app...')
        end_time_hour_min = utils.read_from_file('configs.txt', 'end_time')
        hour = int(end_time_hour_min[0: end_time_hour_min.find('h')])
        minute = int(end_time_hour_min[end_time_hour_min.find('h') + 1: end_time_hour_min.find('m')])
        init = True
        disconnect_dll = False
        load_stocks()
        while True:
            time.sleep(0.1)
            start_time = datetime.datetime.now(tmzone).replace(hour=10, minute=0, second=0, microsecond=0)
            end_time = datetime.datetime.now(tmzone).replace(hour=hour, minute=minute, second=0, microsecond=0)
            while start_time < datetime.datetime.now(tmzone) < end_time:
                if init:
                    nelogica_data_feed_api.dll_initialize()
                    for asset in stocks_list:
                        nelogica_data_feed_api.subscribe_ticker(asset, 'B')
                    disconnect_dll = True
                    init = False
                else:
                    started_at = datetime.datetime.now()
                    now = datetime.datetime.now(tmzone)
                    generate_quotation_excel_sheet(start_date=now - datetime.timedelta(minutes=1), stop_date=now)
                    elapsed_time = (datetime.datetime.now() - started_at).total_seconds()
                    sleep_time = loop_duration - elapsed_time
                    sleep_time = sleep_time if sleep_time >= 0 else 0
                    time.sleep(sleep_time)
            if disconnect_dll:
                nelogica_data_feed_api.dll_disconnect()
                logger.info('App finished!')
                disconnect_dll = False
                nelogica_data_feed_api.dol_player_position = 0
                nelogica_data_feed_api.players_position = {}
    except Exception as e:
        logger.error('Error when running main method with: ' + str(e))


def load_stocks():
    global stocks_list
    configs_file_path = utils.get_dir_path("configs.txt")
    input_data = utils.get_file_as_dict(configs_file_path)
    if input_data is not None and 'assets' in input_data:
        stocks_list = input_data['assets'].split(',')
    print('assets loaded!')


def generate_quotation_excel_sheet(date_str=None, start_date=None, stop_date=None):
    try:
        from openpyxl import Workbook
        if start_date is None or stop_date is None:
            if date_str is not None:
                date = datetime.datetime.strptime(date_str, '%d/%m/%Y').astimezone(tmzone)
                start_date = date.replace(hour=10, minute=0, second=0, microsecond=0)
                stop_date = date.replace(hour=18, minute=0, second=0, microsecond=0)
            else:
                start_date = datetime.datetime.now(tmzone).replace(hour=10, minute=0, second=0, microsecond=0)
                stop_date = datetime.datetime.now(tmzone).replace(hour=18, minute=0, second=0, microsecond=0)
        date = start_date
        open_excel_file_name = 'open_price_' + str(date)[0:10] + '.xlsx'
        close_excel_file_name = 'close_price_' + str(date)[0:10] + '.xlsx'
        high_excel_file_name = 'high_price_' + str(date)[0:10] + '.xlsx'
        low_excel_file_name = 'low_price_' + str(date)[0:10] + '.xlsx'
        volume_excel_file_name = 'volume_' + str(date)[0:10] + '.xlsx'
        if os.path.exists(open_excel_file_name):
            workbook_open = openpyxl.load_workbook(open_excel_file_name)
            first = False
        else:
            workbook_open = Workbook()
            first = True
        if os.path.exists(close_excel_file_name):
            workbook_close = openpyxl.load_workbook(close_excel_file_name)
        else:
            workbook_close = Workbook()
        if os.path.exists(high_excel_file_name):
            workbook_high = openpyxl.load_workbook(high_excel_file_name)
        else:
            workbook_high = Workbook()
        if os.path.exists(low_excel_file_name):
            workbook_low = openpyxl.load_workbook(low_excel_file_name)
        else:
            workbook_low = Workbook()
        if os.path.exists(volume_excel_file_name):
            workbook_volume = openpyxl.load_workbook(volume_excel_file_name)
        else:
            workbook_volume = Workbook()
        while date < stop_date:
            quot_dict = get_quotation_dict(date, date + datetime.timedelta(minutes=1))
            date = date + datetime.timedelta(minutes=1)
            open_sheet = workbook_open.active
            close_sheet = workbook_close.active
            high_sheet = workbook_high.active
            low_sheet = workbook_low.active
            volume_sheet = workbook_volume.active
            header = []
            if first:
                header = list(quot_dict.keys())
            open_values_row = ['' if value is None else value['open'] for key, value in quot_dict.items()]
            close_values_row = ['' if value is None else value['close'] for key, value in quot_dict.items()]
            high_values_row = ['' if value is None else value['max_price'] for key, value in quot_dict.items()]
            low_values_row = ['' if value is None else value['min_price'] for key, value in quot_dict.items()]
            volume_values_row = ['' if value is None else value['volume'] for key, value in quot_dict.items()]

            if first and len(header) > 0:
                header.append('START TIME')
                header.append('STOP TIME')
                open_sheet.append(header)
                close_sheet.append(header)
                high_sheet.append(header)
                low_sheet.append(header)
                volume_sheet.append(header)
                first = False
            if len(open_values_row) > 0:
                open_values_row.append(str(date - datetime.timedelta(minutes=1)))
                open_values_row.append(str(date))
                open_sheet.append(open_values_row)

                close_values_row.append(str(date - datetime.timedelta(minutes=1)))
                close_values_row.append(str(date))
                close_sheet.append(close_values_row)

                high_values_row.append(str(date - datetime.timedelta(minutes=1)))
                high_values_row.append(str(date))
                high_sheet.append(high_values_row)

                low_values_row.append(str(date - datetime.timedelta(minutes=1)))
                low_values_row.append(str(date))
                low_sheet.append(low_values_row)

                volume_values_row.append(str(date - datetime.timedelta(minutes=1)))
                volume_values_row.append(str(date))
                volume_sheet.append(volume_values_row)
            # Save the workbook to a file
            save_workbook(workbook_open, open_excel_file_name)
            save_workbook(workbook_close, close_excel_file_name)
            save_workbook(workbook_high, high_excel_file_name)
            save_workbook(workbook_low, low_excel_file_name)
            save_workbook(workbook_volume, volume_excel_file_name)
    except Exception as e:
        print('Error when running generate_quotation_excel_sheet with: ' + str(e))


def save_workbook(workbook, workbook_path, repeat_count=1):
    try:
        workbook.save(workbook_path)
        print(f'Excel file "{workbook_path}" updated successfully.')
    except Exception:
        if repeat_count <= 5:
            logger.info(f"Trying to save {workbook_path} workbook once more. Count is {repeat_count}")
            time.sleep(repeat_count)
            save_workbook(workbook, workbook_path, repeat_count + 1)
        else:
            logger.warning(f"Exceeded number of tries when saving worksheet {workbook_path}")


def get_quotation_dict(start_date=None, stop_date=None):
    quot_dict = {}
    for stock in stocks_list:
        quotation = get_quotation(stock, start_date, stop_date)
        if quotation is not None:
            quot_dict[stock] = quotation
        else:
            quot_dict[stock] = None
    return quot_dict


def get_quotation(asset_code, start_date=None, stop_date=None):
    trades = None
    with lock:
        if asset_code in nelogica_data_feed_api.trades_dict:
            trades = nelogica_data_feed_api.trades_dict[asset_code].copy()
    if trades is not None and len(trades) > 0:
        df = pd.DataFrame(trades, columns=trade_data_columns)
        df.set_index('trade_date', inplace=True)
        # df.index = pd.to_datetime(df['trade_date'], unit='ms')
        if start_date is not None and stop_date is not None:
            result = {}
            start_date_str = start_date.strftime("%Y-%m-%d %H:%M:%S.%f")
            end_date_str = stop_date.strftime("%Y-%m-%d %H:%M:%S.%f")
            df_selection = df.loc[start_date_str:end_date_str]
            entries = len(df_selection)
            if df_selection is not None and entries > 0:
                # remove processed entries from dict to free up memory
                with lock:
                    nelogica_data_feed_api.trades_dict[asset_code] = nelogica_data_feed_api.trades_dict[asset_code][
                                                                     entries:]
                min_price = df_selection['price'].min()
                max_price = df_selection['price'].max()
                open_price = df_selection.head(1)['price']
                close_price = df_selection.tail(1)['price']
                volume = df_selection['volume'].sum()
                trades = len(df_selection)
                result['min_price'] = min_price
                result['max_price'] = max_price
                result['open'] = open_price
                result['close'] = close_price
                result['volume'] = volume
                result['num_trades'] = trades
                return result


if __name__ == "__main__":
    main()
